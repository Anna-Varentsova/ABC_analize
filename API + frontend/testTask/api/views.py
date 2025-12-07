import openpyxl
import datetime
import requests
import psycopg2
from io import BytesIO
from rest_framework.decorators import api_view
from django.http import HttpResponse
from django.conf import settings
from collections import defaultdict


@api_view(["POST"])
def load_data(request):     #возвращает ответ для фронтенда
    date = request.data.get("date")
    if not date:
        return HttpResponse("Дата не выбрана", status=400)
    now = datetime.datetime.now()
    items = analyze_data(date)
    recording_sql(items, now)
    output = forming_excel(items, date)
    response = HttpResponse(
                            output,
                            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
    response["Content-Disposition"] = f'attachment; filename="ABC_{date}.xlsx"'
    return response


def analyze_data(date):     # возвращает список из сортированных данных
    url = settings.API_URL.format(date=date)
    response = requests.get(url, headers={"Authorization": f"Bearer {settings.API_KEY}"})
    json_data = response.json()  # список из словарей
    summed = defaultdict(lambda: {"sum": 0, "subject": None})
    for row in json_data:
        inc = row["nmId"]
        summed[inc]["subject"] = row["subject"]
        summed[inc]["sum"] += row["finishedPrice"]
    # Преобразуем в список
    items = [{"nmId": k, "subject": v["subject"], "sum": v["sum"]} for k, v in
             summed.items()]
    # Сортировка по сумме
    items.sort(key=lambda x: x["sum"], reverse=True)
    # Подсчёт общей суммы
    total_sum = sum(x["sum"] for x in items)
    # Расчёт долей и ABC-категорий
    cumulative = 0
    for item in items:
        item["share"] = item["sum"] / total_sum * 100
        cumulative += item["share"]
        item["cum_share"] = cumulative
        if item["cum_share"] <= 80:
            item["ABC"] = "A"
        elif item["cum_share"] <= 95:
            item["ABC"] = "B"
        else:
            item["ABC"] = "C"
    return items


def recording_sql(items, now, clear=False):     # записываем в БД
    conn = None
    cur = None
    db = settings.DATABASES['default']
    try:
        conn = psycopg2.connect(
            host=db['HOST'],
            database=db['NAME'],
            user=db['USER'],
            password=db['PASSWORD']
        )
        if conn is not None:
            cur = conn.cursor()
            if not clear:  # если записываем
                sql = ("INSERT INTO \"positionsABC\" ("
                       "\"positionName\","
                       "\"nmId\", "
                       "\"sumTotal\", "
                       "\"sharePercent\", "
                       "\"cumSharePercent\", "
                       "\"ABC\", "
                       "\"queryDateTime\") "
                       "VALUES (%s, %s, %s, %s, %s, %s, %s);")
                for it in items:
                    values = (it['subject'],
                              it['nmId'],
                              it['sum'],
                              it['share'],
                              it['cum_share'],
                              it['ABC'],
                              now)
                    cur.execute(sql, values)
            else:  # если очищаем
                cur.execute("TRUNCATE TABLE \"positionsABC\"")

        conn.commit()
    except (Exception, psycopg2.DatabaseError) as error:
        print("Ошибка:", error)
        print("Тип ошибки:", type(error))
        if conn:
            conn.rollback()  # Откат транзакции в случае ошибки
    finally:
        if cur:
            cur.close()
        if conn:
            conn.close()


def forming_excel(items, date):     #заполняем документ excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"ABC ({date})"
    # Заголовки столбцов (по ключам словаря)
    headers = list(items[0].keys())
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 30
    # Данные
    for row, item in enumerate(items, start=2):
        for col, header in enumerate(headers, start=1):
            value = item.get(header, "")
            cell = ws.cell(row=row, column=col, value=value)
            if col in (4, 5) and isinstance(value, (float, int)):
                cell.number_format = "0.00"
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

